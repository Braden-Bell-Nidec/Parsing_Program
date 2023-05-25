#Import required modules
import pandas as pd
import openpyxl as pyxl
from Combiner import merge_files
from Combiner import delete_temp
import Excel_Functions as ef

DEFAULT_PERCENT = 0.07  #Default outlier percentage threshold

#Set display options for pandas DataFrame outputs in terminal
#pd.set_option('display.max_rows', None) #Debug line
pd.set_option('display.max_colwidth', None)

#Get EPGA data file and Active Directory file names from user respectively.
#Add file extension if user does not include it
EPGA_File = input("Enter path of EPGA file: ")
if not EPGA_File.endswith('.xlsx'):
    EPGA_File += '.xlsx'

AD_File = input("Enter path of Active Directory file: ")
if not AD_File.endswith('.csv'):
    AD_File += '.csv'

#Attempt to combine the given EPGA and Active Directory files into one .xlsx file
try:
    fileName = merge_files(EPGA_File, AD_File, 'combined.xlsx')
except PermissionError: #This *shouldn't* be a common trip because Excel only disallows writing to files while open, but if there's some other program that has it tied up this should catch it.
    print(f"Permission denined when trying to access one of the files {EPGA_File} or {AD_File}")
    print("This is usually cased by one or more of the files being used by another program.")
    print("Please close the program(s) and try again.")
    input("Press enter to close.")
except Exception as e: 
    print(f"Error merging files! Details: {e}")
    input("Press enter to close.")
    exit()

#Attempt to read the temporary combined file into a pandas DataFrame
try:
    df = pd.read_excel(fileName)
except Exception as e: 
    print(f"Error reading the merged Excel file! Details: {e}")
    print("Press enter to close.")
    exit()

#Get custom outlier percentage from user and validate the input
userPercent = input("Enter outlier percentage threshold (default is 7%): ")
if userPercent == "": #If the user just presses enter it uses default case. This is intentional.
    userPercent = DEFAULT_PERCENT #Defined at top of program
else:
    try:
        userPercent = abs(float(userPercent) / 100.00) #Convert the user's percent to decimal format
    except ValueError:
        print("Input value error, resorting to default")
        userPercent = DEFAULT_PERCENT
    except Exception as e:
        print(f"An unknown error occured! Details: {e}")
        print("Resorting to default value and attempting to continue...")
        userPercent = DEFAULT_PERCENT

#Create a dictionary of DataFrames for each unique department in the DataFrame
department_dfs = {}
for department in df['DEPARTMENT'].drop_duplicates().values:
    department_dfs[department] = df[df['DEPARTMENT'] == department]

#Calculate counts and percentages for each department, job title, and responsibility
try:
    #Note: the size() function returns the number of rows in each group
    #reset_index(name='counts') gives a name to the series returned by size()
    counts = df.groupby(['DEPARTMENT', 'JOB_TITLE', 'RESPONSIBILITY_NAME']).size().reset_index(name='counts')
    total_counts = counts.groupby(['DEPARTMENT', 'JOB_TITLE'])['counts'].sum().reset_index(name='total_counts')
    counts = pd.merge(counts, total_counts, on=['DEPARTMENT', 'JOB_TITLE'])
    counts['percentage'] = counts['counts'] / counts['total_counts']
    outliers = counts[counts['percentage'] < userPercent]  #Calculate outliers
except Exception as e:
    print(f"Error calculating counts and percentages! Details: {e}")
    input("Press enter to close.")
    exit()

#Prepare a list of outlier users based on percentage and outlier criteria (user's custom percentage)
outlier_users_list = []
for _, outlier in outliers.iterrows():
    for user in df[
        (df['DEPARTMENT'] == outlier['DEPARTMENT']) & 
        (df['JOB_TITLE'] == outlier['JOB_TITLE']) & 
        (df['RESPONSIBILITY_NAME'] == outlier['RESPONSIBILITY_NAME'])
    ]['USER_NAME']:
        outlier_users_list.append({
            'USER_NAME': user, 
            'DEPARTMENT': outlier['DEPARTMENT'],
            'JOB_TITLE': outlier['JOB_TITLE'], 
            'RESPONSIBILITY_NAME': outlier['RESPONSIBILITY_NAME'],
            'PERCENTAGE': round(outlier['percentage']*100,2)
        })

#Put the outlier user list into a DataFrame
outlier_users = pd.DataFrame(outlier_users_list)
print("\n\n-========================================= Possible Outliers =========================================-")
print(outlier_users)  #Print outlier DataFrame to terminal

#Prepare non-outlier data, i.e., data that does *not* meet the outlier criteria (user's custom percentage)
non_outlier_list = []
for _, row in counts.iterrows():
    if not (
        (outliers['DEPARTMENT'] == row['DEPARTMENT']) & 
        (outliers['JOB_TITLE'] == row['JOB_TITLE']) & 
        (outliers['RESPONSIBILITY_NAME'] == row['RESPONSIBILITY_NAME'])
    ).any():
        non_outlier_list.append({
            'DEPARTMENT': row['DEPARTMENT'],
            'JOB_TITLE': row['JOB_TITLE'], 
            'RESPONSIBILITY_NAME': row['RESPONSIBILITY_NAME'],
            'PERCENTAGE': round(row['percentage']*100,2)
        })

non_outliers = pd.DataFrame(non_outlier_list)
print("\n\n-=========================================== Non-Outliers ===========================================-")
print(non_outliers)  #Print non-outlier DataFrame data to terminal
try:
    #Prepare to write data to an Excel .xlsx
    wb = pyxl.Workbook()  #Create new workbook
    wb.remove(wb.active)  #Remove the default sheet

    #Create and populate the outlier sheet
    outliers_sheet = wb.create_sheet(title='Outliers')
    ef.append_dataframe_to_sheet(outliers_sheet, outlier_users, start_row=1, start_col=1)

    #Format outlier sheet
    ef.adjust_column_width(outliers_sheet, {'A': 15, 'B': 15, 'C': 35, 'D': 42, 'E': 12})
    ef.align_cells(outliers_sheet, ['A', 'B', 'C', 'D'], ef.Alignment(horizontal='center'))

    #Create and populate the non-outlier sheet
    non_outliers_sheet = wb.create_sheet(title='Non-Outliers')
    ef.append_dataframe_to_sheet(non_outliers_sheet, non_outliers, start_row=1, start_col=1)

    #Format non-outlier sheet
    ef.adjust_column_width(non_outliers_sheet, {'A': 15, 'B': 35, 'C': 32, 'D': 12})
    ef.align_cells(non_outliers_sheet, ['A', 'B', 'C', 'D'], ef.Alignment(horizontal='center'))
except Exception as e:
    print(f"Error preparing Excel workbook! Details: {e}")
    input("Press enter to close.")
    exit()

#Create seperate Excel sheets and charts for each unique job title
ef.create_job_title_sheets_and_charts(df, wb)

#Save changes to the workbook
try:
    wb.save("analysis.xlsx")
except PermissionError: #I kept forgetting to close the previous analysis.xlsx file when developing this program so I figured someone else may run into this issue as well at some point
    print("Permission denied when trying to save results to file!")
    print("This is ususally caused by another file called 'analysis.xlsx' being open in Excel (or another program) in the current working directory.")
    print("If a previous analysis.xlsx file is open in Excel (or another program), close the file and save it to another directory if you do not want it to be overwritten!")
    input("Press enter to close.")
    exit()
except Exception as e:
    print(f"Error saving file! Details: {e}")
    delete_temp('combined.xlsx')
    exit()

#Delete the temporary combined file created earlier
delete_temp('combined.xlsx')

#Ending prompt
input("\n\nAll tasks completed. Press enter to close.")