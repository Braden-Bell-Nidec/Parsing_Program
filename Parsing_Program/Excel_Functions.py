from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment
import pandas as pd

def sanitize_sheet_name(sheet_name):
    """
    Sanitizes a sheet name by replacing invalid characters,
    abbreviating certain words, and truncating the name.

    Args:
        sheet_name (str): The original sheet name.

    Returns:
        str: The sanitized sheet name.
    """

    # Define a list of invalid characters
    invalid_chars = ['_', '-', ',', '\\', '/', '*', '[', ']', ':', '?', ' ']

    # Define a dictionary of word abbreviations
    abbreviation_dict = {
        'Manager': 'Mgr',
        'manager': 'Mgr',
        'Manger': 'Mgr',
        'Associate': 'Assoc',
        'I': '1',
        'II': '2',
        'III': '3',
        'InformationTechnology': 'IT',
        'Technician': 'Tech',
        'Mechanical': 'Mech',
        'Certification': 'Cert',
        'Senior': 'Sr.',
        'HumanResources': 'HR',
        'HumanResouce': 'HR',
        'BuisinessPartner': 'BP',
        'President': 'Pres',
        'Engineer': 'Eng',
        'Engineering': 'Eng',
        'Operations': 'Op',
        'MANKATO': 'MKTO',
        'LEXINGTON': 'LEX',
        'REYNOSA': 'REY',
        'Electronic': 'Elec',
        'Component': 'Comp',
        'Assembler': 'Assem',
        'Marketing': 'MkTg',
        'Maintenance': 'Maint.',
        'And': '&',
        'and': '&',
        'General': 'Gen',
        'Network': 'Net',
        'Infrastructure': 'Inf',
        'Specialist': 'Spclst',
        'Environmental': 'Enviro',
        'Health': 'Hlth',
        'Safety': 'Sfty',
        'Director': 'Dir',
        'Administrative': 'Admin',
        'Administrator': 'Admin',
        'Product': 'Prod',
        'Accounts': 'Accts',
        'Aftermarket': 'AM',
        'Remanufacturing': 'Reman',
        'Supervisor': 'Suprvsr.',
        'Development': 'Dev',
        'Caterpillar': 'CAT',
        'Communications': 'Comms',
        'Logistics': 'Log',
        'Compliance': 'Cmpl',
        'Shipping': 'Ship',
        'Recieving': 'Rec',
        'Technical': 'Tech',
        'Fabrication': 'Fab',
        'Manufacturing': 'Man.',
        'Represenative': 'Rep'
    }

    # Replace words in the sheet name with their abbreviations
    for word in sheet_name.split():
        if word in abbreviation_dict:
            sheet_name = sheet_name.replace(word, abbreviation_dict[word])

    # Remove invalid characters from the sheet name
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '')

    # If sheet_name is an empty string after removing invalid chars,
    # replace it with "Unnamed"
    if sheet_name == "":
        sheet_name = "Unnamed"

    # Truncate the sheet name to 31 characters
    sheet_name = sheet_name[:31]

    return sheet_name


def append_dataframe_to_sheet(ws, df, start_row=1, start_col=1):
    """
    Appends a pandas DataFrame to a worksheet. DataFrame values start at the specified row and column.

    Args:
        ws (Worksheet): An openpyxl worksheet object.
        df (DataFrame): A pandas DataFrame.
        start_row (int): The row index where the dataframe should start.
        start_col (int): The column index where the dataframe should start.
    """

    #Write data from the DataFrame to the worksheet
    for i, row in enumerate(df.values, start=1):  #Starting index from 1
        for j, item in enumerate(row, start=1):  #Starting index from 1
            ws.cell(row=start_row+i, column=start_col+j-1, value=item)

    #Write column names to the worksheet
    for j, col_name in enumerate(df.columns, start=1):  #Starting index from 1
        ws.cell(row=start_row+1, column=start_col+j-1, value=col_name)


def create_job_title_sheets_and_charts(df, wb):
    """
    Creates individual sheets and charts for each office location with job title in a DataFrame.

    Args:
        df (pandas.DataFrame): The DataFrame containing job title and office information.
        wb (openpyxl.Workbook): The Workbook object to create sheets and charts.

    Returns:
        None
    """
    # Group by 'OFFICE' and then 'JOB_TITLE'
    grouped = df.groupby(['OFFICE', 'JOB_TITLE'])

    # Iterate over each group
    for (office, title), group in grouped:
        # Sanitize the title and office names
        sanitized_office = sanitize_sheet_name(office)
        sanitized_title = sanitize_sheet_name(title)

        # Combine sanitized office and title names with a hyphen
        combined_title = (sanitized_office + "-" + sanitized_title)[:31]

        # Create a new sheet with the combined title
        ws = wb.create_sheet(title=combined_title)

        # Write the JOB_TITLE to the top of the sheet
        ws.cell(row=1, column=1, value=title)

        # Create pie charts for the office group on the sheet
        create_pie_charts(group, ws)



def adjust_column_width(sheet, cols_width_dict):
    """
    Adjusts the column width of specified columns in a worksheet.

    Args:
        sheet (Worksheet): An openpyxl worksheet object.
        cols_width_dict (dict): A dictionary containing column letters as keys and widths as values.
    """
    #Iterate over each column 
    for col, width in cols_width_dict.items():
        #And set width
        sheet.column_dimensions[col].width = width

def align_cells(sheet, cols, alignment):
    """
    Aligns cells in specified columns to a given alignment.

    Args:
        sheet (Worksheet): An openpyxl worksheet object.
        cols (list): A list of column letters to be aligned.
        alignment (Alignment): An openpyxl Alignment object.
    """
    #Iterate over each column
    for col in cols:
        #iterate over each cell in current column
        for cell in sheet[col]:
            #set alignment of current cell
            cell.alignment = alignment


def create_excel_pie_chart(sheet, df, min_col, max_col, chart_location):
    """
    Creates a pie chart in an Excel sheet based on data from a DataFrame.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The Worksheet object to add the chart to.
        df (pandas.DataFrame): The DataFrame containing the chart data.
        min_col (int): The minimum column index for the chart data.
        max_col (int): The maximum column index for the chart data.
        chart_location (str): The location on the sheet where the chart should be placed.

    Returns:
        None
    """

    # Create a PieChart object
    chart = PieChart()

    # Define the range of labels from the sheet
    labels = Reference(sheet, min_col=min_col, min_row=4, max_row=len(df) + 1)  # Adjusted min_row and max_row

    # Define the range of data values from the sheet
    data = Reference(sheet, min_col=max_col, min_row=4, max_row=len(df) + 1)  # Adjusted min_row and max_row

    # Add the data and set the categories for the chart
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    # Set the title of the chart based on the min_col value
    chart.title = 'Responsibility Distribution' if min_col == 1 else 'Member Of Distribution'

    # Add the chart to the sheet at the specified location
    sheet.add_chart(chart, chart_location)

def create_pie_charts(df, ws):
    """
    Creates pie charts based on data from a DataFrame and adds them to a worksheet.

    Args:
        df (pandas.DataFrame): The DataFrame containing the chart data.
        ws (openpyxl.worksheet.worksheet.Worksheet): The Worksheet object to add the charts to.

    Returns:
        None
    """

    # Extract responsibility and member_of data from the DataFrame
    responsibilities, member_of = get_responsibility_and_member_of_data(df)

    # Append the responsibility data to the worksheet
    append_dataframe_to_sheet(ws, responsibilities, start_row=2, start_col=1)

    # Adjust the column width for better visibility
    adjust_column_width(ws, {'A': 45, 'B': 10})

    # Align cells in column B to center
    align_cells(ws, ['B'], Alignment(horizontal='center'))

    # Create a pie chart for the responsibility data and add it to the worksheet
    create_excel_pie_chart(ws, responsibilities, min_col=1, max_col=2, chart_location="F3")

    # Append the member_of data to the worksheet at a column offset
    append_dataframe_to_sheet(ws, member_of, start_row=2, start_col=4)  # 4 here is an offset for a new column

    # Adjust the column width for better visibility
    adjust_column_width(ws, {'D': 45, 'E': 20})

    # Align cells in column E to center
    align_cells(ws, ['E'], Alignment(horizontal='center'))

    # Create a pie chart for the member_of data and add it to the worksheet
    create_excel_pie_chart(ws, member_of, min_col=4, max_col=5, chart_location="F18")

def split_and_explode(df, column, delimiter=';'):
    """
    Splits the values in a column based on a delimiter and returns a DataFrame where each row contains a single value from the split data.

    Args:
        df (pandas.DataFrame): The DataFrame containing the column to split.
        column (str): The column to split.
        delimiter (str): The delimiter to use for splitting the column. Defaults to ';'.

    Returns:
        pandas.DataFrame: A DataFrame where each row contains a single value from the split data.
    """

    # Split the values in the column based on the delimiter
    s = df[column].str.split(delimiter).apply(pd.Series, 1).stack()

    # Reset the index of the resulting Series
    s.index = s.index.droplevel(-1)

    # Set the name of the resulting Series to the original column name
    s.name = column

    # Drop the original column from the DataFrame and join the split data as a new column
    return df.drop(columns=column).join(s)



def get_responsibility_and_member_of_data(df):
    """
    Retrieves responsibility and 'Member of' data from a given DataFrame.

    Args:
        df (pandas.DataFrame): A DataFrame to retrieve responsibility and 'Member of' data from.

    Returns:
        tuple: A tuple containing two DataFrames: (responsibilities, member_of).

    Raises:
        KeyError: If 'RESPONSIBILITY_NAME' or 'MEMBER_OF' column doesn't exist in the DataFrame.
    """

    # Get the counts of each responsibility name in the DataFrame
    responsibilities = df['RESPONSIBILITY_NAME'].value_counts().reset_index()
    responsibilities.columns = ['RESPONSIBILITY_NAME', 'COUNTS']

    # Split and explode the 'MEMBER_OF' column in the DataFrame
    member_of_df = split_and_explode(df, 'MEMBER_OF')

    # Get the counts of each 'Member of' value in the exploded DataFrame
    member_of = member_of_df['MEMBER_OF'].value_counts().reset_index()
    member_of.columns = ['MEMBER_OF', 'COUNTS_MEMBER_OF']

    return responsibilities, member_of