#Import required modules
import pandas as pd
import openpyxl as pyxl
import Excel_Functions as ef
import tkinter as tk
from time import sleep
from GUI import GUI
from sys import exit
from Combiner import merge_files
from Combiner import delete_temp

def main(EPGA_File, AD_File, user_percent, delete_combined, progress, status):
    """
    The main function that performs the analysis on EPGA and Active Directory files.

    Args:
        EPGA_File (str): The name of the EPGA Excel file.
        AD_File (str): The name of the Active Directory CSV file.
        user_percent (str): The outlier percentage threshold specified by the user.
        delete_combined (bool): Indicates whether to delete the temporary combined file.
        progress (tkinter.ttk.Progressbar): The progress bar widget to update the progress.
        status (tkinter.StringVar): The status message variable to update the status.

    Returns:
        None
    """

    DEFAULT_PERCENT = 0.07  # Default outlier percentage threshold

    # Set display options for pandas DataFrame outputs in terminal
    pd.set_option('display.max_colwidth', None)

    # Get EPGA data file and Active Directory file names from user respectively.
    # Add file extension if user does not include it

    # Attempt to combine the given EPGA and Active Directory files into one .xlsx file
    status.set("Merging files...")
    try:
        fileName = merge_files(EPGA_File, AD_File, 'combined.xlsx')
        progress['value'] = 10

    except PermissionError:
        print(f"Permission denied when trying to access one of the files {EPGA_File} or {AD_File}")
        print("This is usually caused by one or more of the files being used by another program.")
        print("Please close the program(s) and try again.")
        status.set("Error")
        exit()

    except UnicodeDecodeError:
        print("Unicode decode error. Make sure the correct files were selected and are formatted properly.")
        status.set("Error")
        exit()

    except FileNotFoundError:
        print("One or more of the files could not be found.")
        status.set("Error")
        exit()

    # Attempt to read the temporary combined file into a pandas DataFrame
    try:
        status.set("Reading combined file...")
        df = pd.read_excel(fileName)
        progress['value'] = 20
    except Exception as e:
        status.set("Error")
        print(f"Error reading the merged Excel file! Details: {e}")
        exit()

    # Validate and process the user-defined outlier percentage
    if user_percent == "":
        user_percent = DEFAULT_PERCENT
    else:
        try:
            if int(user_percent) < 100:
                user_percent = abs(float(user_percent) / 100.00)
            else:
                print("Percent was greater than 100.\nResorting to default.")
                user_percent = DEFAULT_PERCENT
        except ValueError:
            status.set("Error")
            print("Input value error, resorting to default")
            user_percent = DEFAULT_PERCENT
        except Exception as e:
            status.set("Error")
            print(f"An unknown error occurred! Details: {e}")
            print("Resorting to default value and attempting to continue...")
            user_percent = DEFAULT_PERCENT

    # Create a dictionary of DataFrames for each unique department in the DataFrame
    department_dfs = {}
    for department in df['DEPARTMENT'].drop_duplicates().values:
        department_dfs[department] = df[df['DEPARTMENT'] == department]

    progress['value'] = 30
    status.set("Calculating percentages...")

    try:
        # Calculate counts and percentages for each department, job title, and responsibility
        counts = df.groupby(['DEPARTMENT', 'JOB_TITLE', 'RESPONSIBILITY_NAME']).size().reset_index(name='counts')
        total_counts = counts.groupby(['DEPARTMENT', 'JOB_TITLE'])['counts'].sum().reset_index(name='total_counts')
        counts = pd.merge(counts, total_counts, on=['DEPARTMENT', 'JOB_TITLE'])
        counts['percentage'] = counts['counts'] / counts['total_counts']
        outliers = counts[counts['percentage'] < user_percent]
    except Exception as e:
        status.set("Error")
        print(f"Error calculating counts and percentages! Details: {e}")
        exit()

    progress['value'] = 40

    outlier_users_list = []
    for _, outlier in outliers.iterrows():
        # Iterate through the DataFrame to identify outlier users
        for user in df[
            (df['DEPARTMENT'] == outlier['DEPARTMENT']) &
            (df['JOB_TITLE'] == outlier['JOB_TITLE']) &
            (df['RESPONSIBILITY_NAME'] == outlier['RESPONSIBILITY_NAME'])
        ]['USER_NAME']:
            outlier_users_list.append({
                'USER_NAME': user,
                'DEPARTMENT': outlier['DEPARTMENT'],
                'JOB_TITLE': outlier['JOB_TITLE'],
                'RESPONSIBILITY_NAME': outlier['RESPONSIBILITY_NAME'],
                'PERCENTAGE': round(outlier['percentage']*100, 2)
            })

    outlier_users = pd.DataFrame(outlier_users_list)
    progress['value'] = 50
    print("-========================================= Possible Outliers =========================================-")
    print(outlier_users)

    non_outlier_list = []
    for _, row in counts.iterrows():
        # Iterate through the counts DataFrame to identify non-outliers
        if not (
            (outliers['DEPARTMENT'] == row['DEPARTMENT']) &
            (outliers['JOB_TITLE'] == row['JOB_TITLE']) &
            (outliers['RESPONSIBILITY_NAME'] == row['RESPONSIBILITY_NAME'])
        ).any():
            non_outlier_list.append({
                'DEPARTMENT': row['DEPARTMENT'],
                'JOB_TITLE': row['JOB_TITLE'],
                'RESPONSIBILITY_NAME': row['RESPONSIBILITY_NAME'],
                'PERCENTAGE': round(row['percentage']*100, 2)
            })

    non_outliers = pd.DataFrame(non_outlier_list)
    print("\n\n-=========================================== Non-Outliers ===========================================-")
    print(non_outliers)

    progress['value'] = 60
    status.set("Preparing Excel workbook...")

    try:
        wb = pyxl.Workbook()
        wb.remove(wb.active)

        outliers_sheet = wb.create_sheet(title='Outliers')
        ef.append_dataframe_to_sheet(outliers_sheet, outlier_users, start_row=1, start_col=1)

        ef.adjust_column_width(outliers_sheet, {'A': 15, 'B': 15, 'C': 35, 'D': 42, 'E': 12})
        ef.align_cells(outliers_sheet, ['A', 'B', 'C', 'D'], ef.Alignment(horizontal='center'))

        non_outliers_sheet = wb.create_sheet(title='Non-Outliers')
        ef.append_dataframe_to_sheet(non_outliers_sheet, non_outliers, start_row=1, start_col=1)

        ef.adjust_column_width(non_outliers_sheet, {'A': 15, 'B': 35, 'C': 32, 'D': 12})
        ef.align_cells(non_outliers_sheet, ['A', 'B', 'C', 'D'], ef.Alignment(horizontal='center'))
    except Exception as e:
        status.set("Error")
        print(f"Error preparing Excel workbook! Details: {e}")
        exit()

    # Create separate Excel sheets and charts for each unique job title
    ef.create_job_title_sheets_and_charts(df, wb)
    status.set("Saving Excel workbook...")
    progress['value'] = 70

    try:
        wb.save("analysis.xlsx")
    except PermissionError:
        status.set("Error")
        print("Permission denied when trying to save results to file!")
        print("This is usually caused by another file called 'analysis.xlsx' being open in Excel (or another program) in the current working directory.")
        exit()
    except Exception as e:
        status.set("Error")
        print(f"Error saving file! Details: {e}")
        ef.delete_temp('combined.xlsx')
        exit()

    progress['value'] = 80

    if delete_combined:
        status.set("Cleaning up...")
        delete_temp('combined.xlsx')
        sleep(0.25)

    progress['value'] = 100
    status.set("All tasks completed.")
   
#Run the GUI
root = tk.Tk()
gui = GUI(root, main)
root.mainloop()